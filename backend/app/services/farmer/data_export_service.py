# backend/app/services/farmer/data_export_service.py

from datetime import datetime
from typing import Dict, Any, Optional, List
import json
import csv
import io
import base64

# Excel support
from openpyxl import Workbook

# PDF support
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4


# ----------------------------
# Internal imports (safe, in-memory)
# ----------------------------
try:
    from app.services.farmer.unit_service import _unit_store
except:
    _unit_store = {}

try:
    from app.services.farmer.season_calendar_service import _calendar_store
except:
    _calendar_store = {}

try:
    from app.services.farmer.task_service import _task_templates_store
except:
    _task_templates_store = {}

try:
    from app.services.farmer.financial_ledger_service import _ledger_store
except:
    _ledger_store = []

try:
    from app.services.farmer.input_shortage_service import _input_inventory_store
except:
    _input_inventory_store = {}

try:
    from app.services.farmer.irrigation_service import get_irrigation_schedule
except:
    get_irrigation_schedule = None

try:
    from app.services.farmer.recommendation_engine_service import generate_recommendations_for_unit
except:
    generate_recommendations_for_unit = None


# ============================================================
# Helpers — collect all relevant farmer data per unit
# ============================================================
def collect_farmer_data(unit_id: Optional[str] = None) -> Dict[str, Any]:
    """
    Create a full export dictionary containing:
     - Units
     - Calendar
     - Tasks
     - Ledger
     - Inventory
     - Irrigation
     - Recommendations
    """

    data = {
        "exported_at": datetime.utcnow().isoformat(),
        "units": {},
        "calendar": {},
        "task_templates": _task_templates_store,
        "ledger": list(_ledger_store),
        "inventory": _input_inventory_store,
        "irrigation": {},
        "recommendations": {},
    }

    units = [unit_id] if unit_id else list(_unit_store.keys())
    for uid in units:
        unit = _unit_store.get(uid)
        if unit:
            data["units"][uid] = unit

        cal = _calendar_store.get(uid)
        if cal:
            data["calendar"][uid] = cal

        if get_irrigation_schedule:
            try:
                sch = get_irrigation_schedule(uid)
                if sch:
                    data["irrigation"][uid] = sch
            except:
                pass

        if generate_recommendations_for_unit:
            try:
                rec = generate_recommendations_for_unit(uid)
                data["recommendations"][uid] = rec
            except:
                pass

    return data


# ============================================================
# JSON export
# ============================================================
def export_json(unit_id: Optional[str] = None) -> str:
    data = collect_farmer_data(unit_id)
    return json.dumps(data, indent=2)


# ============================================================
# CSV export (multi-sheet concept flattened)
# ============================================================
def export_csv(unit_id: Optional[str] = None) -> Dict[str, str]:
    """
    Returns dict:
      {
        "units.csv": <csv string>,
        "ledger.csv": <csv string>,
        "inventory.csv": <csv string>,
        "calendar.csv": <csv string>
      }
    """

    data = collect_farmer_data(unit_id)

    result = {}

    # Units CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["unit_id", "name", "crop", "area", "stage_template_id"])
    for uid, u in data["units"].items():
        writer.writerow([uid, u.get("name"), u.get("crop"), u.get("area"), u.get("stage_template_id")])
    result["units.csv"] = buf.getvalue()

    # Ledger CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["entry_id", "type", "category", "amount", "date", "description"])
    for e in data["ledger"]:
        writer.writerow([e.get("entry_id"), e.get("type"), e.get("category"), e.get("amount"), e.get("date"), e.get("description")])
    result["ledger.csv"] = buf.getvalue()

    # Inventory CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["item_id", "name", "quantity", "unit", "min_threshold"])
    for iid, r in data["inventory"].items():
        writer.writerow([iid, r.get("name"), r.get("quantity"), r.get("unit"), r.get("min_threshold")])
    result["inventory.csv"] = buf.getvalue()

    # Calendar CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["unit_id", "task_name", "stage_name", "scheduled_start", "scheduled_end"])
    for uid, cal in data["calendar"].items():
        for e in cal.get("entries", []):
            writer.writerow([
                uid,
                e.get("task_name"),
                e.get("stage_name"),
                e.get("scheduled_start_iso"),
                e.get("scheduled_end_iso"),
            ])
    result["calendar.csv"] = buf.getvalue()

    return result


# ============================================================
# Excel export (XLSX)
# ============================================================
def export_xlsx(unit_id: Optional[str] = None) -> bytes:
    data = collect_farmer_data(unit_id)
    wb = Workbook()

    # Units sheet
    ws = wb.active
    ws.title = "Units"
    ws.append(["unit_id", "name", "crop", "area", "stage_template"])
    for uid, u in data["units"].items():
        ws.append([uid, u.get("name"), u.get("crop"), u.get("area"), u.get("stage_template_id")])

    # Ledger
    ws = wb.create_sheet("Ledger")
    ws.append(["entry_id", "type", "category", "amount", "date", "description"])
    for e in data["ledger"]:
        ws.append([
            e.get("entry_id"),
            e.get("type"),
            e.get("category"),
            e.get("amount"),
            e.get("date"),
            e.get("description"),
        ])

    # Inventory
    ws = wb.create_sheet("Inventory")
    ws.append(["item_id", "name", "quantity", "unit", "min_threshold"])
    for iid, r in data["inventory"].items():
        ws.append([iid, r.get("name"), r.get("quantity"), r.get("unit"), r.get("min_threshold")])

    # Calendar
    ws = wb.create_sheet("Calendar")
    ws.append(["unit_id", "task_name", "stage_name", "start", "end"])
    for uid, cal in data["calendar"].items():
        for e in cal.get("entries", []):
            ws.append([
                uid,
                e.get("task_name"),
                e.get("stage_name"),
                e.get("scheduled_start_iso"),
                e.get("scheduled_end_iso"),
            ])

    # Recommendations (summary only)
    ws = wb.create_sheet("Recommendations")
    ws.append(["unit_id", "category", "text", "severity", "score"])
    for uid, recs in data["recommendations"].items():
        for r in recs.get("recommendations", []):
            ws.append([
                uid,
                r.get("category"),
                r.get("recommendation"),
                r.get("severity"),
                r.get("score"),
            ])

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ============================================================
# PDF export (summary report)
# ============================================================
def export_pdf(unit_id: Optional[str] = None) -> bytes:
    data = collect_farmer_data(unit_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    story = []

    story.append(Paragraph("<b>Farmer Report</b>", styles["Title"]))
    story.append(Paragraph(f"Generated at {datetime.utcnow().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Units summary
    story.append(Paragraph("<b>Production Units</b>", styles["Heading2"]))
    for uid, u in data["units"].items():
        story.append(Paragraph(f"Unit {uid}: {u.get('name')} — Crop: {u.get('crop')} — Area: {u.get('area')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    # Inventory summary
    story.append(Paragraph("<b>Inventory</b>", styles["Heading2"]))
    for iid, r in data["inventory"].items():
        story.append(Paragraph(f"{iid}: {r.get('name')} — {r.get('quantity')} {r.get('unit')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    # Ledger summary
    story.append(Paragraph("<b>Ledger Summary</b>", styles["Heading2"]))
    story.append(Paragraph(f"Total entries: {len(data['ledger'])}", styles["Normal"]))
    story.append(Spacer(1, 10))

    doc.build(story)
    return buffer.getvalue()


# ============================================================
# Unified export handler
# ============================================================
def export_data(format: str, unit_id: Optional[str] = None) -> Dict[str, Any]:
    format = format.lower()

    if format == "json":
        return {"format": "json", "data": export_json(unit_id)}

    if format == "csv":
        return {"format": "csv", "files": export_csv(unit_id)}

    if format == "xlsx":
        b = export_xlsx(unit_id)
        encoded = base64.b64encode(b).decode()
        return {"format": "xlsx", "base64": encoded}

    if format == "pdf":
        b = export_pdf(unit_id)
        encoded = base64.b64encode(b).decode()
        return {"format": "pdf", "base64": encoded}

    return {"error": "unsupported_format"}
